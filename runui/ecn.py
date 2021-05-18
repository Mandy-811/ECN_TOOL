# —*-  coding:utf-8  -*-
"""
作者：Mandy.Meng
日期：2021年04月20日
"""
import sys
from ECN import Ui_ECN
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5 import QtWidgets
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import subprocess
import os.path
import time


class ECN(QtWidgets.QWidget, Ui_ECN):
    def __init__(self):
        super(ECN, self).__init__()
        self.setupUi(self)
        self.bom = None
        self.record = None
        self.ecn = None
        self.pushButton_openbom.clicked.connect(self.open_bom)
        self.pushButton_openrecord.clicked.connect(self.open_record)
        self.pushButton_openECN.clicked.connect(self.open_ecn)
        self.pushButton_start.clicked.connect(self.start)

    def open_bom(self):
        filedialog = QFileDialog()
        filedialog.setFileMode(QFileDialog.AnyFile)
        smt_row = None
        nm_row = None
        if filedialog.exec():
            filenames = filedialog.selectedFiles()
            # 获取的文件名添加到LineEdit控件显示
            self.bom = filenames[0]
            self.lineEdit_bompath.setText(self.bom)
            try:
                excel = load_workbook(self.bom)
                table = excel[excel.sheetnames[1]]
                for col in list(table.columns)[0]:
                    if col.value == "SMT Parts":
                        smt_row = col.row
                    if col.value == "Not Mounted":
                        nm_row = col.row
            except:
                self.pushButton_ImportInquire.setEnabled(False)
                QMessageBox.warning(self, '警告', "Error: BOM文件格式错误", QMessageBox.Yes |
                                    QMessageBox.No, QMessageBox.Yes)
            else:
                subprocess.Popen(self.bom, shell=True, encoding='utf-8')
        return

    def open_record(self):
        self.lineEdit_recordpath.setText("变更记录路径")
        return

    def open_ecn(self):
        self.lineEdit_ECNpath.setText("ecn路径")
        return

    def start(self):
        print("开始")
        return







if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainwindow = ECN()
    mainwindow.show()
    sys.exit(app.exec_())
